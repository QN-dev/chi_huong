from flask import Flask,render_template,request,redirect,url_for
import openpyxl as xl
from datetime import datetime,timedelta
from collections import OrderedDict
import os
from utlis import *

app = Flask(__name__)
app.debug = True



@app.route('/')
def index():
    auto_check(90)
    wb,sheet,max_row,max_col=open_file('data/company_info.xlsx',active=True)
    data=[]
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=get_column_index(sheet,'status')).value=='to_call':
            temp={}
            temp=get_data_from_sheet(sheet,i)
            data.append(temp)
    return render_template('index.html',data=data)


@app.route('/insert', methods=['GET', 'POST'])
def insert():
    if request.method=='GET':
        return render_template('insert.html')
    elif request.method == 'POST':
        data={}
        data['business_name']=request.form.get('business_name')
        data['agent']=request.form.get('agent')
        data['possition']=request.form.get('possition')
        data['address']=request.form.get('address')
        data['phone']=request.form.get('phone')
        data['contact_person']=request.form.get('contact_person')
        data['contact_phone']=request.form.get('contact_phone')
        data['fax']=request.form.get('fax')
        data['tax_number']=request.form.get('tax_number')
        data['bank_account']=request.form.get('bank_account')
        data['bank_name']=request.form.get('bank_name')
        data['contact_value']=request.form.get('contact_value')
        data['status']=request.form.get('status')
        data['date_to_call']=request.form.get('date_to_call')
        if len(data)>0:
            insert_company_info('data/company_info.xlsx',data)
            return redirect(url_for('listing'))
        else:
            return render_template('messange.html',message='Vui lòng điền thông tin')

@app.route('/edit/<int:company_id>', methods=['GET', 'POST'])
def edit(company_id):
    if request.method=='GET':
        wb,sheet,max_row,max_col=open_file('data/company_info.xlsx',active=True)
        for i in range(2,max_row+1):
            if sheet.cell(row=i,column=get_column_index(sheet,'id')).value == company_id:
                data = get_data_from_sheet(sheet,i)
        return render_template('edit.html',data=data)
    elif request.method == 'POST':
        data={}
        data['id']=company_id
        data['business_name']=request.form.get('business_name')
        data['agent']=request.form.get('agent')
        data['possition']=request.form.get('possition')
        data['address']=request.form.get('address')
        data['phone']=request.form.get('phone')
        data['contact_person']=request.form.get('contact_person')
        data['contact_phone']=request.form.get('contact_phone')
        data['fax']=request.form.get('fax')
        data['tax_number']=request.form.get('tax_number')
        data['bank_account']=request.form.get('bank_account')
        data['bank_name']=request.form.get('bank_name')
        data['contact_value']=request.form.get('contact_value')
        data['status']=request.form.get('status')
        data['date_to_call']=request.form.get('date_to_call')
        print('postted new data')
        if len(data)>0:
            edit_company_info('data/company_info.xlsx',data)
            return redirect(url_for('company_info',company_id=company_id))
        else:
            return render_template('messange.html',message='Vui lòng điền thông tin')

@app.route('/listing')
def listing():
    wb,sheet,max_row,max_col=open_file('data/company_info.xlsx',active=True)
    data=[]
    for i in range(2,max_row+1):
        temp={}
        temp=get_data_from_sheet(sheet,i)
        data.append(temp)
    return render_template('company_list.html',data=data)


@app.route('/company/<int:company_id>',methods=['GET','POST'])
def company_info(company_id):
    '''
    data's order:
    'business_name'
    'agent'
    'possition'
    'address'
    'phone'
    'fax'
    'tax_number'
    'bank_account'
    'bank_name'
    'contact_value'
    'status'
    'date_to_call'

    history's order:
    'id'
    'company_id'
    'method':mail/call
    'time'
    'change_status'
    'note'
    'status'
    '''
    # XXX: get note data
    wb_note,sheet_note,max_row_note,max_col_note=open_file('data/history.xlsx',active=True)
    note=[]
    for i in range(1,max_row_note+1):
        if sheet_note.cell(row=i,column=get_column_index(sheet_note,'company_id')).value == company_id:#just get note of that company
            temp = get_note_from_sheet(sheet_note,i)
            note.append(temp)

    # XXX: get company info data
    wb,sheet,max_row,max_col=open_file('data/company_info.xlsx',active=True)
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=get_column_index(sheet,'id')).value == company_id:
            data = get_data_from_sheet(sheet,i)


    if request.method=='GET':
        return render_template('company_info.html',data=data,notes=note,company_id=company_id)
    elif request.method=='POST':
        # BUG: haven't have method mail yet
        get_status=request.form.get('status')
        is_change = True if get_status != data['status'] else False
        history={}
        history['company_id']=data['id']
        history['business_name']=data['business_name']
        history['method']='call'
        history['time']=datetime.now().strftime("%d-%m-%Y %H:%M")
        history['note']=request.form.get('note')
        history['change_status']=is_change
        history['status']= request.form.get('status')
        write_history('data/history.xlsx',history)
        #write company status if has change in status
        if is_change:
            write_change('data/company_info.xlsx',data['id'],'status',request.form.get('status'))
            print(data['status'],'and ',request.form.get('status'))

            #auto update when done calling
            if request.form.get('status') =='done' :
                time_to_call= datetime.strptime(data['date_to_call'],'%d-%m-%Y') + timedelta(days=365*2)
                data['date_to_call'] =  datetime.strftime(time_to_call,'%d-%m-%Y')
                write_change('data/company_info.xlsx',data['id'],'date_to_call',data['date_to_call'])
        else:
            pass
        return redirect('/company/'+str(company_id))


@app.route('/history')
def history():
    '''
    data's order:
    'business_name'
    'time'
    'note'
    'status'
    '''
    wb,sheet,max_row,max_col=open_file('data/history.xlsx',active=True)
    data=[]
    for i in range(2,max_row+1):
        temp=get_note_from_sheet(sheet,i)
        data.append(temp)

    return render_template('history.html',data=data)
