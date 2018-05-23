import openpyxl as xl
from datetime import datetime,timedelta





def open_file(file_path,active=False):# XXX: create a saving file
    try:
        wb=xl.load_workbook(file_path)
    except:
        wb=xl.Workbook()
        wb.save(file_path)
        wb=xl.load_workbook(file_path)

    if active:
        sheet=wb.active
        return wb,sheet,sheet.max_row,sheet.max_column # NOTE: max_row is the current row has written
    else:
        return wb


def get_column_index(sheet,entry):
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1,column=i).value == entry:
            return i
    return None


def auto_check(number_of_day_to_announce):
    now=datetime.now()
    wb,sheet,max_row,max_col=open_file('data/company_info.xlsx',active=True)
    for i in range(2,max_row+1):
        status = sheet.cell(row=i,column=get_column_index(sheet,'status')).value
        if status != 'to_call' :#just check the status is not tocall
            date = sheet.cell(row=i,column=get_column_index(sheet,'date_to_call')).value
            if  now + timedelta(days=number_of_day_to_announce) > datetime.strptime(date,'%d-%m-%Y'):
                sheet.cell(row=i,column=get_column_index(sheet,'status')).value='to_call'
    wb.save('data/company_info.xlsx')



def insert_company_info(save_path,data):
    '''
    data's order:
    'business_name'
    'agent'
    'possition'
    'address'
    'phone'
    'fax'
    'contact_person'
    'contact_phone'
    'tax_number'
    'bank_account'
    'bank_name'
    'contact_value'
    'status'
    'date_to_call'
    '''
    wb,sheet,max_row,max_col=open_file(save_path,active=True)
    id = sheet.cell(row=max_row,column=1).value+1 if str(sheet.cell(row=max_row,column=1).value).isdigit() else 0
    row_to_write=max_row+1
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'id')).value=id
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'business_name')).value=data['business_name']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'agent')).value=data['agent']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'possition')).value=data['possition']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'address')).value=data['address']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'phone')).value=data['phone']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'fax')).value=data['fax']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'tax_number')).value=data['tax_number']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'contact_person')).value=data['contact_person']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'contact_phone')).value=data['contact_phone']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'bank_account')).value=data['bank_account']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'bank_name')).value=data['bank_name']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'contact_value')).value=data['contact_value']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'status')).value=data['status']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'date_to_call')).value=data['date_to_call']
    print('Wrote new data')
    wb.save(save_path)

def edit_company_info(save_path,data):
    '''
    data's order:
    'business_name'
    'agent'
    'possition'
    'address'
    'phone'
    'fax'
    'contact_person'
    'contact_phone'
    'tax_number'
    'bank_account'
    'bank_name'
    'contact_value'
    'status'
    'date_to_call'
    '''
    wb,sheet,max_row,max_col=open_file(save_path,active=True)
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=get_column_index(sheet,'id')).value == data['id']:
            row_to_write = i
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'id')).value=data['id']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'business_name')).value=data['business_name']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'agent')).value=data['agent']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'possition')).value=data['possition']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'address')).value=data['address']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'phone')).value=data['phone']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'fax')).value=data['fax']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'tax_number')).value=data['tax_number']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'contact_person')).value=data['contact_person']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'contact_phone')).value=data['contact_phone']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'bank_account')).value=data['bank_account']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'bank_name')).value=data['bank_name']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'contact_value')).value=data['contact_value']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'status')).value=data['status']
            sheet.cell(row=row_to_write,column=get_column_index(sheet,'date_to_call')).value=data['date_to_call']
    print('Edited data')
    wb.save(save_path)


def write_change(data_path,id,column,value):
    wb,sheet,max_row,max_col=open_file(data_path,active=True)
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=get_column_index(sheet,'id')).value==id:
            sheet.cell(row=i,column=get_column_index(sheet,column)).value=value
    wb.save(data_path)


def get_data_from_sheet(sheet,row):
    '''
    'business_name'
    'agent'
    'possition'
    'address'
    'phone'
    'fax'
    'tax_number'
    'contact_person'
    'contact_phone'
    'bank_account'
    'bank_name'
    'contact_value'
    'status'
    'status_write'
    'date_to_call'
    '''
    data={}
    data['id']=sheet.cell(row=row,column=get_column_index(sheet,'id')).value
    data['business_name']=sheet.cell(row=row,column=get_column_index(sheet,'business_name')).value
    data['agent']=sheet.cell(row=row,column=get_column_index(sheet,'agent')).value
    data['possition']=sheet.cell(row=row,column=get_column_index(sheet,'possition')).value
    data['address']=sheet.cell(row=row,column=get_column_index(sheet,'address')).value
    data['phone']=sheet.cell(row=row,column=get_column_index(sheet,'phone')).value
    data['contact_person']=sheet.cell(row=row,column=get_column_index(sheet,'contact_person')).value
    data['contact_phone']=sheet.cell(row=row,column=get_column_index(sheet,'contact_phone')).value
    data['fax']=sheet.cell(row=row,column=get_column_index(sheet,'fax')).value
    data['tax_number']=sheet.cell(row=row,column=get_column_index(sheet,'tax_number')).value
    data['bank_account']=sheet.cell(row=row,column=get_column_index(sheet,'bank_account')).value
    data['bank_name']=sheet.cell(row=row,column=get_column_index(sheet,'bank_name')).value
    contact_value = sheet.cell(row=row,column=get_column_index(sheet,'contact_value')).value
    data['contact_value']='{:0,}đ'.format(int(contact_value)) if contact_value is not None else 'Không có dữ liệu'
    data['status']=sheet.cell(row=row,column=get_column_index(sheet,'status')).value
    data['date_to_call']=sheet.cell(row=row,column=get_column_index(sheet,'date_to_call')).value
    if data['status'] == 'to_call':
        data['status_write']= 'Cần gọi'
    elif data['status'] == 'calling':
        data['status_write'] ='Đang gọi'
    else:
        data['status_write']='Đã gọi'
    return data


def get_note_from_sheet(sheet,row):
    '''
    'business_name'
    'agent'
    'possition'
    'address'
    'phone'
    'contact_person'
    'contact_phone'
    'fax'
    'tax_number'
    'bank_account'
    'bank_name'
    'contact_value'
    'status'
    'status_write'
    'date_to_call'
    '''
    data={}
    data['id']=sheet.cell(row=row,column=get_column_index(sheet,'id')).value
    data['time']=sheet.cell(row=row,column=get_column_index(sheet,'time')).value
    data['business_name']=sheet.cell(row=row,column=get_column_index(sheet,'business_name')).value
    note = sheet.cell(row=row,column=get_column_index(sheet,'note')).value if sheet.cell(row=row,column=get_column_index(sheet,'note')).value is not None else 'Chuyển trạng thái sang đã gọi'
    data['note']= note
    data['status']=sheet.cell(row=row,column=get_column_index(sheet,'status')).value
    if data['status'] == 'to_call':
        data['status_write']= 'Cần gọi'
    elif data['status'] == 'calling':
        data['status_write'] ='Đang gọi'
    else:
        data['status_write']='Đã gọi'
    return data


def write_history(save_path,data):
    '''
    history's order:
    'id'
    'company_id'
    'business_name'
    'method':mail/call
    'time'
    'change_status'
    'note'
    'status'
    '''
    wb,sheet,max_row,max_col=open_file(save_path,active=True)
    id = sheet.cell(row=max_row,column=1).value+1 if str(sheet.cell(row=max_row,column=1).value).isdigit() else 0
    row_to_write=max_row+1
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'id')).value=id
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'company_id')).value=data['company_id']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'business_name')).value=data['business_name']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'method')).value=data['method']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'time')).value=data['time']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'change_status')).value=data['change_status']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'note')).value=data['note']
    sheet.cell(row=row_to_write,column=get_column_index(sheet,'status')).value=data['status']
    print('Writing new history')
    wb.save(save_path)
